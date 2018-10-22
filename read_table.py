from openpyxl import load_workbook, Workbook 
src_data = "ExportPageQueryEntry20181021100414.xlsx"

def tchong(li,num):
	if len(li)<num:
		li.append("")
		return tchong(li,num)
	else:
		return li


wb = load_workbook(src_data)
ws = wb["sheet1"]

def read_row(row,dic):
	query = row[1].value
	flag = row[3].value
	content = row[2].value
	# id_ = row[0].value
	# operator = row[4].value
	try:
		dic[query]
	except:
		dic[query] = {}
	try:
		dic[query][flag]
	except:
		dic[query][flag] = []
	dic[query][flag].append(content)


dic = {}
for row in ws[2:len(ws["A"])]:
	read_row(row,dic)

wb_new = Workbook()
ws_new = wb_new.active


row_id = 1
for kw, content in dic.items():
	ws_new[row_id][0].value = kw
	content["Url"] = content["Url"] if "Url" in content else []
	content["AS"] = content["AS"] if "AS" in content else []
	content["QS"] = content["QS"] if "QS" in content else []

	row_num = max(len(content["Url"]), len(content["AS"]), len(content["QS"]))
	for tag,ar in content.items():
		content[tag] = tchong(ar,row_num)	
		# print(tag, ar)
		# print(content[tag])

	for url,AS,QS in zip(content["Url"],content["AS"],content["QS"]):
		ws_new.cell(row = row_id, column = 2).value = url
		ws_new.cell(row = row_id, column = 4).value = AS
		ws_new.cell(row = row_id, column = 5).value = QS
		row_id += 1

		# a,b,c = url, AS, QS
		# print(a,b,c)
		
	
wb_new.save("1.xlsx")



# print(ws[])
